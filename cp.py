import openpyxl
import csv
import datetime
from os import rename
from email.mime.application import MIMEApplication
import smtplib
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from openpyxl.styles import Alignment
import logging
'''
logfilename = './log/test.log'
logging.basicConfig(filename='./log/test.log', level=logging.DEBUG)
logger = logging.getLogger("crumbs")
logger.setLevel(logging.DEBUG)

logging.debug('debug')
logging.info('info')
logging.warning('warning')
logging.error('error')
logging.critical('critical')
'''
'''
# ./log/test.log 결과
DEBUG:root:debug
INFO:root:info
WARNING:root:warning
ERROR:root:error
CRITICAL:root:critical
'''
formatter = logging.Formatter('[%(levlename)s|%(filename)s:%(lineno)s]%(asctime)s>%(message)s')
fileHandler = logging.FileHandler('./log/my.log')

fileHandler.setFormatter(formatter)

logger.debug("debug")
logger.info("info")
logger.warning("warning")
logger.error("error")
logger.critical("critical")

start = datetime.datetime.now()
print('시작시간 : ' + str(start))

f = open('C:/Users/nana/Desktop/conveni.csv','r',encoding='UTF-8')
rdr = csv.reader(f)
cnt = 0
title = []
companytit = []
for a in rdr:
 if cnt ==0:
  for i in range(len(a)):
   title.append(a[i])
   cnt+=1
 else:
  companytit.append(a[4])
supplier = []
for c in companytit:
 if c in supplier:
  continue
 else:
  supplier.append(c)

f = open('C:/Users/nana/Desktop/conveni.csv','r',encoding='UTF-8')
csv = csv.reader(f)
origin = []
dt = datetime.datetime.now()
d = str(dt.date()).split('-')
year = d[0]
month = d[1]
day = d[2]

# supplier 리스트를 회사명으로 변경 필요
# 회사명:지급자명 최종 리스트 불러오기
maplist = openpyxl.load_workbook('c:/Users/nana/Desktop/입점사리스트.xlsx')
worksheet = maplist.active
max = worksheet.max_row
mappinglist = {}
company_name_list = []
ceo_name_list = []
ceo_email_list = []
for i in range(max):
 if worksheet.cell(row=i+2, column=2).value is None:
  continue
 else:
  company_name_list.append(worksheet.cell(row=i+2, column=1).value) #회사이름 저장
  ceo_name_list.append(worksheet.cell(row=i+2, column=2).value) #회사사장이름 저장
  ceo_email_list.append(worksheet.cell(row=i+2, column=3).value) #email 주소 저장

# for i in range(len(nlist)):
# print(str(nlist[i]) + ' : ' + str(clist[i]))
# 열간격 width=20으로 저장할것

# 열간격 수정 20200329
column_list_width = [17,8.5,20,20,9,12,15,95,15,6,9,15,15,90,45,15]
column_list_name = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']
titlecell_list_name = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1']


for c in csv:
 origin.append(c)
for i in range(len(supplier)):
 wb = openpyxl.Workbook()
 sheet = wb.active
 sheet.append(title)
 wb.save('컨비니_{}_{}_{}_{}_발주서.xlsx'.format(year,month,day,supplier[i]))
 list = []
 for j in range(len(origin)):
  if origin[j][4] == supplier[i]:
   load_wb = openpyxl.load_workbook('컨비니_{}_{}_{}_{}_발주서.xlsx'.format(year,month,day,supplier[i]))
   load_ws = load_wb['Sheet']
   #sheet = load_ws.active
   #list.append(origin[j])
   print(origin[j])
   sheet.append(origin[j])
   for k in range(len(column_list_name)):
    sheet.column_dimensions[column_list_name[k]].width = column_list_width[k] # 20200329 간격 처리
   for m in range(len(titlecell_list_name)): #가운데 정렬
    cell = sheet.cell(row=1,column=m+1)
    cell.alignment = Alignment(horizontal='center') #가운데 정렬 끝
   #sheet.column_dimensions['C'].width  20.5
   #sheet.column_dimensions['D'].width = 15
   wb.save('컨비니_{}_{}_{}_{}_발주서.xlsx'.format(year,month,day,supplier[i]))
  else:
   continue
company_final_list = []
#파일명 변경 프로세스
for i in range(len(supplier)): #공급자명 리스트 for 문
 for j in range(len(ceo_name_list)): #ceo_name_list for 문
  if supplier[i] == ceo_name_list[j]: #공급자명과 ceo_name_list 비교
   company_final_list.append(company_name_list[j])
   #print('공급자명 : ' + supplier[i] + ', ' + '대표명 : ' + ceo_name_list[j] + ', ' + '회사명 : ' + company_name_list[j])
   try :
    rename('C:/Users/nana/Desktop/컨비니_{}_{}_{}_{}_발주서.xlsx'.format(year, month, day, ceo_name_list[j]),'컨비니_{}_{}_{}_{}_발주서.xlsx'.format(year, month, day, company_name_list[j]))
   except:
    print(supplier[i] + ' 또는 ' + company_name_list[j] + '에 일치하는 회사 또는 대표명이 없습니다.')
  else:
   continue

#filenamelist = company_final_list

#for i in range(10):
 #sending_email('africa352@naver.com', filenamelist[i])

end = datetime.datetime.now()

timediff = end - start

print('소요시간 : ' + str(timediff))

#f.close()
